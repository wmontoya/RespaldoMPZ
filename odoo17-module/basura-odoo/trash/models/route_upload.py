from odoo import models, fields, api, _
from odoo.exceptions import UserError
import xml.etree.ElementTree as ET
import base64
import re
import io
from datetime import datetime
import openpyxl
import unicodedata
import math

class RouteUpload(models.TransientModel):
    _name = "trash.route_upload"
    _description = "Upload KML and Excel Files"

    kml_file = fields.Binary(string="KML File (Rutas)", required=False)
    filename = fields.Char(string="KML Filename")
    
    excel_file = fields.Binary(string="Excel File (INFO 1)", required=False)
    excel_filename = fields.Char(string="Excel Filename")

    excel_data = fields.Text(string="Excel Data (Internal)", store=False)

    def _get_default_sector(self):
        Sector = self.env['trash.route_sector'].sudo()
        sector = Sector.search([('name_sector', '=', 'SIN SECTOR')], limit=1)
        if not sector:
            sector = Sector.create({'name_sector': 'SIN SECTOR'})
        return sector.id

    def _get_or_create_sector(self, sector_name):
        if not sector_name or not sector_name.strip():
            return self._get_default_sector()
        
        sector_name = sector_name.strip().upper()
        Sector = self.env['trash.route_sector'].sudo()
        sector = Sector.search([('name_sector', '=', sector_name)], limit=1)
        if not sector:
            sector = Sector.create({'name_sector': sector_name})
        return sector.id

    def kml_color_to_hex(self, kml_color):
        if not kml_color or len(kml_color) != 8:
            return '#000000'

        bb = int(kml_color[2:4], 16)
        gg = int(kml_color[4:6], 16)
        rr = int(kml_color[6:8], 16)
        hex_color = f"#{rr:02x}{gg:02x}{bb:02x}"

        return hex_color

    def process_coordinates(self, coords_text):
        coords_list = []
        if not coords_text:
            return coords_list
            
        raw_coords = coords_text.strip()
        raw_coords = re.sub(r'\s+', ' ', raw_coords)
        
        coord_parts = raw_coords.strip().split()
        
        for part in coord_parts:
            if not part:
                continue
                
            components = part.split(',')
            if len(components) >= 2:
                try:
                    lon = float(components[0].strip())
                    lat = float(components[1].strip())
                    
                    if (-86 <= lon <= -82) and (8 <= lat <= 11):
                        coords_list.append((lat, lon))
                except (ValueError, IndexError):
                    continue
        
        return coords_list

    def split_coordinates_by_distance(self, coords_list, max_gap_km=0.5):
        if not coords_list or len(coords_list) < 2:
            return [coords_list] if coords_list else []
        
        def haversine_distance(lat1, lon1, lat2, lon2):
            R = 6371
            lat1_rad = math.radians(lat1)
            lat2_rad = math.radians(lat2)
            delta_lat = math.radians(lat2 - lat1)
            delta_lon = math.radians(lon2 - lon1)
            
            a = math.sin(delta_lat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(delta_lon/2)**2
            c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
            return R * c
        
        segments = []
        current_segment = [coords_list[0]]
        
        for i in range(1, len(coords_list)):
            prev_lat, prev_lon = coords_list[i-1]
            curr_lat, curr_lon = coords_list[i]
            
            distance = haversine_distance(prev_lat, prev_lon, curr_lat, curr_lon)
            
            if distance > max_gap_km:
                if len(current_segment) >= 2:
                    segments.append(current_segment)
                current_segment = [coords_list[i]]
            else:
                current_segment.append(coords_list[i])
        
        if len(current_segment) >= 2:
            segments.append(current_segment)
        
        return segments

    def _normalize_route_code(self, code):
        if code is None:
            return ''
        code_str = str(code).strip()
        code_str = code_str.replace('\xa0', ' ')
        code_str = re.sub(r'\s+', ' ', code_str)
        return code_str.strip()
    
    def _normalize_header(self, header):
        if header is None:
            return ''
        header_str = str(header)
        header_str = unicodedata.normalize('NFD', header_str)
        header_str = ''.join(c for c in header_str if unicodedata.category(c) != 'Mn')
        header_str = header_str.replace('\n', ' ').replace('\r', ' ')
        header_str = re.sub(r'\s+', ' ', header_str)
        return header_str.strip().upper()

    def extract_days(self, text):
        if not text:
            return []
        
        text = str(text).upper()
        
        text = re.sub(r'\([^)]*\)', '', text)
        
        for extra in ["CADA 15 DÍAS", "CADA 8 DÍAS", "SEMANAL", "QUINCENAL", "MENSUAL", "CADA SEMANA"]:
            text = text.replace(extra, "")
        
        day_names = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO", "SABADO", "DOMINGO"]
        
        day_variants = {
            "LUNES": 0, "MARTES": 1, "MIÉRCOLES": 2, "MIERCOLES": 2,
            "JUEVES": 3, "VIERNES": 4, "SÁBADO": 5, "SABADO": 5, "DOMINGO": 6
        }
        
        text_normalized = text.replace('É', 'E').replace('Á', 'A').replace('Ó', 'O')
        
        range_match = re.search(r'([A-ZÁÉÍÓÚÑ]+)\s+A\s+([A-ZÁÉÍÓÚÑ]+)', text)
        if range_match:
            start_day = range_match.group(1).upper()
            end_day = range_match.group(2).upper()
            
            start_norm = start_day.replace('É', 'E').replace('Á', 'A')
            end_norm = end_day.replace('É', 'E').replace('Á', 'A')
            
            start_idx = -1
            end_idx = -1
            
            for day, idx in day_variants.items():
                day_norm = day.replace('É', 'E').replace('Á', 'A')
                if day_norm == start_norm or day == start_day:
                    start_idx = idx
                if day_norm == end_norm or day == end_day:
                    end_idx = idx
            
            if start_idx >= 0 and end_idx >= 0 and start_idx <= end_idx:
                result_days = day_names[start_idx:end_idx+1]
                return result_days
        
        found_days = []
        
        for day in day_names:
            day_norm = day.replace('É', 'E').replace('Á', 'A')
            pattern = r'\b' + re.escape(day) + r'\b'
            pattern_norm = r'\b' + re.escape(day_norm) + r'\b'
            
            if re.search(pattern, text) or re.search(pattern_norm, text_normalized):
                found_days.append(day)
        
        if not found_days:
            words = re.split(r'[,;Y\s]+', text)
            for word in words:
                word_clean = word.replace('É', 'E').replace('Á', 'A').strip()
                for day, idx in day_variants.items():
                    day_norm = day.replace('É', 'E').replace('Á', 'A')
                    if day_norm == word_clean:
                        found_days.append(day)
                        break
        
        return list(set(found_days))
    
    def extract_collection_time(self, text):
        if not text:
            return 'DIURNA'
        
        text = text.upper()
        
        if 'NOCTURNO' in text:
            return 'NOCTURNA'
        elif 'DIURNO' in text:
            return 'DIURNA'
        
        return 'DIURNA'

    def detect_iteration(self, text):
        if not text:
            return "SEMANAL"
        
        text = text.upper()
        
        if 'QUINCENAL' in text or 'CADA 15' in text or '15 DÍAS' in text or '15 DIAS' in text or 'QUINCENA' in text:
            return "QUINCENAL"
        
        if 'MENSUAL' in text or 'CADA 30' in text or '30 DÍAS' in text or '30 DIAS' in text or 'UNA VEZ AL MES' in text:
            return "MENSUAL"
        
        if 'SEMANAL' in text or 'CADA SEMANA' in text or 'SEMANALMENTE' in text:
            return "SEMANAL"
        
        return "SEMANAL"

    def read_excel_data(self):
        if not self.excel_file:
            return {}
        
        if not openpyxl:
            raise ImportError("openpyxl library is required to read Excel files. Please install it.")
        
        raw_data = self.excel_file
        
        if isinstance(raw_data, str):
            data = base64.b64decode(raw_data)
        elif isinstance(raw_data, bytes):
            if raw_data[:2] == b'PK':
                data = raw_data
            else:
                data = base64.b64decode(raw_data)
        else:
            raise ValueError(f"Tipo de dato de Excel no reconocido: {type(raw_data)}")

        if len(data) < 4:
            raise ValueError("El archivo está vacío o está corrupto.")

        file_signature = data[:4]
        if file_signature != b'PK\x03\x04':
            if file_signature.startswith(b'PK'):
                raise ValueError("El archivo parece ser un archivo ZIP corrupto. Por favor guarde el archivo como .xlsx desde Excel.")
            elif file_signature.startswith(b'\xd0\xcf\x11\xe0'):
                raise ValueError("El archivo está en formato antiguo de Excel (.xls). Por favor guardelo como .xlsx desde Excel.")
            else:
                sig_hex = file_signature.hex()
                raise ValueError(f"El archivo no es un archivo Excel válido. Firma del archivo: {sig_hex}. Por favor verifique que el archivo sea .xlsx y pueda abrirlo en Excel.")
        
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(data))
            except Exception as e2:
                raise ValueError(f"No se pudo abrir el archivo Excel: {str(e2)}")
        ws = wb.active
        
        merged_ranges = ws.merged_cells.ranges
        
        max_col = ws.max_column
        
        headers = None
        header_row = 1
        
        for row_num in range(1, 6):
            row_headers = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                row_headers.append(cell.value)
            
            non_none_count = sum(1 for h in row_headers if h is not None)
            if non_none_count > 3:
                headers = row_headers
                header_row = row_num
                break
        
        if headers is None or all(h is None for h in headers):
            raise ValueError("No se encontraron encabezados en las primeras 5 filas del archivo Excel.")
        
        excel_routes = {}
        rows_processed = 0
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
            rows_processed += 1
            if not row[0]:
                continue
            
            row_data = {}
            raw_row_values = list(row)
            for col_idx, header in enumerate(headers):
                if col_idx < len(row):
                    row_data[header] = row[col_idx]
                    if header:
                        normalized_h = self._normalize_header(header)
                        row_data[normalized_h] = row[col_idx]
            
            normalized_headers = {}
            for idx, h in enumerate(headers):
                if h:
                    normalized_h = self._normalize_header(h)
                    normalized_headers[normalized_h] = idx
                    normalized_headers[str(h).strip() if h else ''] = idx
            
            route_col = None
            for key in ['RUTA', 'ruta', 'Ruta', 'CODIGO RUTA', 'Código Ruta', 'Código', 'Código de Ruta']:
                if key in normalized_headers:
                    route_col = key
                    break
            
            sector_col = None
            for key in ['Sector', 'SECTOR', 'sector', 'NOMBRE SECTOR', 'Nombre Sector']:
                if key in normalized_headers:
                    sector_col = key
                    break
            
            district_col = None
            for key in ['DISTRITO', 'Distrito', 'distrito']:
                if key in normalized_headers:
                    district_col = key
                    break
            
            dates_col = None
            for key in ['FECHAS', 'Fechas', 'FECHAS DE RECICLAJE', 'FECHAS RECICLAJE']:
                if key in normalized_headers:
                    dates_col = key
                    break
            
            route_code = None
            if route_col:
                route_code = row_data.get(route_col)
            
            sector_name = None
            if sector_col:
                sector_name = row_data.get(sector_col)
            
            if not route_code:
                continue
            
            route_code = self._normalize_route_code(route_code)
            
            if route_code not in excel_routes:
                excel_routes[route_code] = {
                    'sectors': {},
                    'collection_types': {},
                    'quincenal_dates': []
                }
            
            if sector_name:
                sector_name = str(sector_name).strip()
                if sector_name not in excel_routes[route_code]['sectors']:
                    excel_routes[route_code]['sectors'][sector_name] = {
                        'waste_types': {},
                        'collection_time': 'DIURNA',
                        'quincenal_dates': []
                    }
                
                waste_type_map_dia = {
                    'BASURA DIA': 'BASURA',
                    'ORGANICO DIA': 'ORGÁNICO',
                    'RECICLAJE DIA': 'RECICLAJE',
                    'NO APROVECHABLE DIA': 'NO APROVECHABLE',
                }
                
                for col_name, waste_type in waste_type_map_dia.items():
                    col_value = row_data.get(col_name)
                    if col_value and str(col_value).strip():
                        col_value_str = str(col_value).strip()
                        
                        if isinstance(col_value, datetime):
                            continue
                        if '-' in col_value_str and any(c.isdigit() for c in col_value_str):
                            continue
                        
                        if waste_type not in excel_routes[route_code]['sectors'][sector_name]['waste_types']:
                            excel_routes[route_code]['sectors'][sector_name]['waste_types'][waste_type] = {
                                'months': [],
                                'days': [],
                                'iteration': 'SEMANAL'
                            }
                        
                        days = self.extract_days(col_value_str)
                        if days:
                            excel_routes[route_code]['sectors'][sector_name]['waste_types'][waste_type]['days'] = days
                        
                        collection_time = self.extract_collection_time(col_value_str)
                        if collection_time:
                            excel_routes[route_code]['sectors'][sector_name]['collection_time'] = collection_time
                        
                        current_iteration = excel_routes[route_code]['sectors'][sector_name]['waste_types'][waste_type]['iteration']
                        new_iteration = self.detect_iteration(col_value_str)
                        
                        if new_iteration != 'SEMANAL' or current_iteration == 'SEMANAL':
                            excel_routes[route_code]['sectors'][sector_name]['waste_types'][waste_type]['iteration'] = new_iteration
                
                for waste_type in ['BASURA', 'ORGÁNICO', 'RECICLAJE', 'NO APROVECHABLE']:
                    for month in ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 
                                  'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']:
                        col_name = f"{waste_type}_{month}"
                        col_value = row_data.get(col_name) or row_data.get(f"{waste_type} {month}")
                        if col_value and str(col_value).strip():
                            if waste_type not in excel_routes[route_code]['sectors'][sector_name]['waste_types']:
                                excel_routes[route_code]['sectors'][sector_name]['waste_types'][waste_type] = {
                                    'months': [],
                                    'days': [],
                                    'iteration': 'SEMANAL'
                                }
                            excel_routes[route_code]['sectors'][sector_name]['waste_types'][waste_type]['months'].append(month)
                
                iteration_col = (
                    row_data.get('ITERACION') or 
                    row_data.get('FRECUENCIA') or
                    row_data.get('FRECUENCIA RECOLECCION') or
                    row_data.get('TIPO') or
                    None
                )
                
                if iteration_col:
                    iteration = self.detect_iteration(str(iteration_col))
                    for waste_type in excel_routes[route_code]['sectors'][sector_name]['waste_types']:
                        current = excel_routes[route_code]['sectors'][sector_name]['waste_types'][waste_type]['iteration']
                        if current == 'SEMANAL':
                            excel_routes[route_code]['sectors'][sector_name]['waste_types'][waste_type]['iteration'] = iteration
                
                time_col = (
                    row_data.get('TURNO') or 
                    row_data.get('HORA') or 
                    row_data.get('HORARIO')
                )
                
                if not time_col:
                    for col_name, waste_type in waste_type_map_dia.items():
                        col_value = row_data.get(col_name)
                        if col_value and str(col_value).strip():
                            time_val = self.extract_collection_time(str(col_value))
                            if time_val and time_val != 'DIURNA':
                                time_col = col_value
                                break
                
                if time_col:
                    time_val = str(time_col).upper()
                    if 'NOCTURNO' in time_val:
                        excel_routes[route_code]['sectors'][sector_name]['collection_time'] = 'NOCTURNA'
                
                reciclaje_iteration = excel_routes[route_code]['sectors'][sector_name]['waste_types'].get('RECICLAJE', {}).get('iteration', 'SEMANAL')
                
                if reciclaje_iteration == 'QUINCENAL':
                    for col_idx in range(7, len(raw_row_values)):
                        date_val = raw_row_values[col_idx]
                        if date_val:
                            try:
                                if isinstance(date_val, datetime):
                                    excel_routes[route_code]['sectors'][sector_name]['quincenal_dates'].append(date_val)
                                elif isinstance(date_val, str) and date_val.strip() and date_val.strip() != '-':
                                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%y']:
                                        try:
                                            parsed = datetime.strptime(date_val.strip(), fmt)
                                            excel_routes[route_code]['sectors'][sector_name]['quincenal_dates'].append(parsed)
                                            break
                                        except:
                                            continue
                            except Exception as e:
                                pass
                    
                    fechas_col = row_data.get('FECHAS DE RECICLAJE')
                    if fechas_col and isinstance(fechas_col, str):
                        date_parts = fechas_col.split('-')
                        for part in date_parts:
                            part = part.strip()
                            if part:
                                for fmt in ['%d/%m/%y', '%d/%m/%Y']:
                                    try:
                                        parsed = datetime.strptime(part, fmt)
                                        if parsed not in excel_routes[route_code]['sectors'][sector_name]['quincenal_dates']:
                                            excel_routes[route_code]['sectors'][sector_name]['quincenal_dates'].append(parsed)
                                        break
                                    except:
                                        continue
        
        return excel_routes

    def process_kml_and_excel(self):
        excel_routes = {}
        if self.excel_file:
            try:
                excel_routes = self.read_excel_data()
            except Exception as e:
                import traceback
                raise UserError(f"Error al leer el archivo Excel: {str(e)}")
        
        DAY_MAP = {
            "LUNES": "monday",
            "MARTES": "tuesday",
            "MIÉRCOLES": "wednesday",
            "JUEVES": "thursday",
            "VIERNES": "friday",
            "SÁBADO": "saturday",
            "DOMINGO": "sunday",
        }
        
        routes_created = 0
        routes_updated = 0
        route_days_created = 0
        route_days_updated = 0
        
        if self.kml_file:
            data = base64.b64decode(self.kml_file)
            root = ET.fromstring(data)
            ns = {'kml': 'http://www.opengis.net/kml/2.2'}

            styles = {}
            for style in root.findall(".//kml:Style", ns):
                style_id = style.get('id')
                color_elem = style.find('kml:LineStyle/kml:color', ns)
                width_elem = style.find('kml:LineStyle/kml:width', ns)

                kml_color = color_elem.text if color_elem is not None else None

                styles[f"#{style_id}"] = {
                    'color': self.kml_color_to_hex(kml_color) if kml_color else None,
                    'width': width_elem.text if width_elem is not None else None
                }

            for placemark in root.findall(".//kml:Placemark", ns):
                styleUrl = placemark.find('kml:styleUrl', ns)
                styleline = styleUrl.text if styleUrl is not None else ''
                style_data = styles.get(styleline + '-normal', {})

                route_data = {}
                extended = placemark.find('kml:ExtendedData', ns)
                if extended is not None:
                    for data_elem in extended.findall('kml:Data', ns):
                        key = data_elem.get('name')
                        value_elem = data_elem.find('kml:value', ns)
                        value = value_elem.text if value_elem is not None else ''
                        route_data[key] = value

                all_line_strings = []
                
                multi_geometry = placemark.find('kml:MultiGeometry', ns)
                if multi_geometry is not None:
                    for linestring in multi_geometry.findall('kml:LineString', ns):
                        coords_elem = linestring.find('kml:coordinates', ns)
                        if coords_elem is not None and coords_elem.text:
                            segment_coords = self.process_coordinates(coords_elem.text)
                            if segment_coords and len(segment_coords) >= 2:
                                split_segments = self.split_coordinates_by_distance(segment_coords)
                                all_line_strings.extend(split_segments)
                else:
                    for linestring in placemark.findall('kml:LineString', ns):
                        coords_elem = linestring.find('kml:coordinates', ns)
                        if coords_elem is not None and coords_elem.text:
                            segment_coords = self.process_coordinates(coords_elem.text)
                            if segment_coords and len(segment_coords) >= 2:
                                split_segments = self.split_coordinates_by_distance(segment_coords)
                                all_line_strings.extend(split_segments)

                if not all_line_strings:
                    continue
                    
                route_code = route_data.get('RUTA')
                route_name = route_data.get('RUTA') or ''
                route_color = style_data.get('color') or ''
                
                route = self.env['trash.route'].search(
                    [('code', '=', route_code)],
                    limit=1
                )

                if route:
                    routes_updated += 1
                    route.write({
                        'name': route_name or route.name,
                        'color': route_color or route.color,
                    })
                else:
                    routes_created += 1
                    route = self.env['trash.route'].create({
                        'name': route_name,
                        'code': route_code or '',
                        'color': route_color,
                    })

                for segment_coords in all_line_strings:
                    segment = self.env['trash.route_segment'].create({
                        'route_id': route.id,
                    })

                    for lat, lon in segment_coords:
                        self.env['trash.route_point'].create({
                            'route_id': route.id,
                            'segment_id': segment.id,
                            'latitude': lat,
                            'longitude': lon,
                        })
                
                normalized_route_code = self._normalize_route_code(route_code)
                route_excel_data = excel_routes.get(normalized_route_code, {})
                
                day_vals = []
                
                if route_excel_data and 'sectors' in route_excel_data:
                    sectors_data = route_excel_data['sectors']
                    
                    for sector_name, sector_info in sectors_data.items():
                        sector_id = self._get_or_create_sector(sector_name)
                        collection_time = sector_info.get('collection_time', 'DIURNA')
                        quincenal_dates = sector_info.get('quincenal_dates', [])
                        
                        waste_types_data = sector_info.get('waste_types', {})
                        
                        for waste_type, type_info in waste_types_data.items():
                            extracted_days = type_info.get('days', [])
                            iteration = type_info.get('iteration', 'SEMANAL')
                            
                            day_ids = []
                            for d in extracted_days:
                                key = DAY_MAP.get(d.upper())
                                if key:
                                    day_rec = self.env['trash.day'].search([('name', '=', key)], limit=1)
                                    if not day_rec:
                                        day_rec = self.env['trash.day'].create({'name': key})
                                    day_ids.append(day_rec.id)
                            
                            if day_ids:
                                existing_route_day = self.env['trash.route_day'].search([
                                    ('route_id', '=', route.id),
                                    ('sector_id', '=', sector_id),
                                    ('waste_type', '=', waste_type),
                                ], limit=1)
                                
                                if existing_route_day:
                                    route_days_updated += 1
                                    existing_route_day.write({
                                        'day_iteration': iteration,
                                        'day_ids': [(6, 0, day_ids)],
                                        'collection_time': collection_time,
                                    })
                                    route_day_rec = existing_route_day
                                else:
                                    route_days_created += 1
                                    route_day_rec = self.env['trash.route_day'].create({
                                        'route_id': route.id,
                                        'waste_type': waste_type,
                                        'day_iteration': iteration,
                                        'day_ids': [(6, 0, day_ids)],
                                        'sector_id': sector_id,
                                        'collection_time': collection_time,
                                    })
                                day_vals.append(route_day_rec.id)
                                
                                if iteration == 'QUINCENAL' and quincenal_dates:
                                    for q_date in quincenal_dates:
                                        date_val = q_date.date() if isinstance(q_date, datetime) else q_date
                                        existing_date = self.env['trash.route_quincenal_date'].search([
                                            ('route_day_id', '=', route_day_rec.id),
                                            ('collection_date', '=', date_val),
                                        ], limit=1)
                                        
                                        if not existing_date:
                                            self.env['trash.route_quincenal_date'].create({
                                                'route_day_id': route_day_rec.id,
                                                'collection_date': date_val,
                                            })

                if day_vals:
                    route.write({'day_ids': [(6, 0, day_vals)]})
        
        elif excel_routes and not self.kml_file:
            for route_code, route_excel_data in excel_routes.items():
                route = self.env['trash.route'].search([
                    ('code', '=', route_code)
                ], limit=1)
                
                if not route:
                    continue
                
                routes_updated += 1
                
                if 'sectors' in route_excel_data:
                    sectors_data = route_excel_data['sectors']
                    
                    for sector_name, sector_info in sectors_data.items():
                        sector_id = self._get_or_create_sector(sector_name)
                        collection_time = sector_info.get('collection_time', 'DIURNA')
                        quincenal_dates = sector_info.get('quincenal_dates', [])
                        
                        waste_types_data = sector_info.get('waste_types', {})
                        
                        for waste_type, type_info in waste_types_data.items():
                            extracted_days = type_info.get('days', [])
                            iteration = type_info.get('iteration', 'SEMANAL')
                            
                            day_ids = []
                            for d in extracted_days:
                                key = DAY_MAP.get(d.upper())
                                if key:
                                    day_rec = self.env['trash.day'].search([('name', '=', key)], limit=1)
                                    if not day_rec:
                                        day_rec = self.env['trash.day'].create({'name': key})
                                    day_ids.append(day_rec.id)
                            
                            if day_ids:
                                existing_route_day = self.env['trash.route_day'].search([
                                    ('route_id', '=', route.id),
                                    ('sector_id', '=', sector_id),
                                    ('waste_type', '=', waste_type),
                                ], limit=1)
                                
                                if existing_route_day:
                                    route_days_updated += 1
                                    existing_route_day.write({
                                        'day_iteration': iteration,
                                        'day_ids': [(6, 0, day_ids)],
                                        'collection_time': collection_time,
                                    })
                                    route_day_rec = existing_route_day
                                else:
                                    route_days_created += 1
                                    route_day_rec = self.env['trash.route_day'].create({
                                        'route_id': route.id,
                                        'waste_type': waste_type,
                                        'day_iteration': iteration,
                                        'day_ids': [(6, 0, day_ids)],
                                        'sector_id': sector_id,
                                        'collection_time': collection_time,
                                    })
                                
                                if iteration == 'QUINCENAL' and quincenal_dates:
                                    for q_date in quincenal_dates:
                                        date_val = q_date.date() if isinstance(q_date, datetime) else q_date
                                        existing_date = self.env['trash.route_quincenal_date'].search([
                                            ('route_day_id', '=', route_day_rec.id),
                                            ('collection_date', '=', date_val),
                                        ], limit=1)
                                        
                                        if not existing_date:
                                            self.env['trash.route_quincenal_date'].create({
                                                'route_day_id': route_day_rec.id,
                                                'collection_date': date_val,
                                            })

        message = f"Se procesaron {routes_created} rutas nuevas y {routes_updated} rutas actualizadas."
        if route_days_created or route_days_updated:
            message += f" Días de ruta: {route_days_created} creados, {route_days_updated} actualizados."
        if self.excel_file:
            message += f" Se cargó información de {len(excel_routes)} rutas desde Excel."
        
        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
            'params': {
                'title': 'Carga de Rutas',
                'message': message,
            }
        }

    def process_kml(self):
        return self.process_kml_and_excel()
